import openpyxl

outputWorkbook = openpyxl.Workbook()
outputSheet = outputWorkbook.active


def process_xlsx(filename, row_number, col_segments):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    min_col = 1
    index = 1
    row_index = 2
    cropType = ""

    columnNames = ["cropName", "cropType", "KcIni", "KcMid", "KcEnd"]

    for i in range(1, 6):
        outputCell = outputSheet.cell(row=1, column=i)
        outputCell.value = columnNames[i-1]

    # Process rows for the first 5 columns until the specified row number
    while index <= col_segments:
        for row in sheet.iter_rows(min_row=3, max_row=row_number, min_col=min_col, max_col=min_col + 4):
            if row[0].value == None:
                continue

            cell_index = 1
            while cell_index < 6:

                if row[3].value == "Kc mid":
                    cropType = row[1].value
                    print(cropType)
                    row_index -= 1
                    break

                if cell_index == 1:
                    outputCell = outputSheet.cell(
                        row=row_index, column=cell_index)

                    value = row[cell_index].value

                    if row[cell_index].value[-1] == "*":
                        value = row[cell_index].value[:-1]

                    outputCell.value = value.strip()
                    cell_index += 1
                    continue

                elif cell_index == 2:
                    outputCell = outputSheet.cell(
                        row=row_index, column=cell_index)
                    outputCell.value = cropType
                    cell_index += 1

                outputCell = outputSheet.cell(row=row_index, column=cell_index)
                outputCell.value = row[cell_index-1].value
                cell_index += 1

            row_index += 1
        index += 1
        min_col += 5


# Usage example
process_xlsx('Crop_coefficients_values.xlsx', 39, 4)
outputWorkbook.save('processed_crop_coefficients_values.xlsx')


"""
changes made to Crop_coefficients_values.xlsx:

added space between * and cropName in cell B9 for consistency in processing

"""
